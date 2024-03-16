import google.generativeai as genai
import pandas as pd
import regex as re
import openpyxl

original = []
translated = []

class translator:
    language = None
    add_on_prompt = None

    def __init__(self) :

        self.gen_model()

    def read_file(self,file):

        workbook = pd.ExcelFile(file)
        sheet_names = workbook.sheet_names
        return sheet_names
    
    def excel_to_df(self,file,sheet_name):

        self.df = pd.read_excel(file,sheet_name=sheet_name)
        
    def gen_model(self):

        genai.configure(api_key="AIzaSyBD5XH3iDNycU2Q4HeRYXi4z_9oAjXMYWc")
        safety_settings = [
            {
                "category": "HARM_CATEGORY_DANGEROUS",
                "threshold": "BLOCK_NONE",
            },
            {
                "category": "HARM_CATEGORY_HARASSMENT",
                "threshold": "BLOCK_NONE",
            },
            {
                "category": "HARM_CATEGORY_HATE_SPEECH",
                "threshold": "BLOCK_NONE",
            },
            {
                "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
                "threshold": "BLOCK_NONE",
            },
            {
                "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
                "threshold": "BLOCK_NONE",
            },
        ]
        self.model = genai.GenerativeModel('gemini-1.0-pro-latest', safety_settings = safety_settings)


    def translate(self,cell_value,prompt=None):

        if not pd.isna(cell_value):
            
            if prompt is None:
                gen_ai_prompt= f"""
                Translate the context below into {self.language}. Ensure that you do not elaborate too much and do not miss any points.
                Do translation keeping in mind the following points:{self.add_on_prompt}
                context= {cell_value}
                """
            else:
                gen_ai_prompt = f"{prompt} context= {cell_value}"
            
            response = self.model.generate_content(gen_ai_prompt)
            return response.text
        else:
            return cell_value
        
    def column_index(self,column_name):

        index = 0
        for char in column_name:
            index = index * 26 + (ord(char) - ord('A')) + 1
        return index - 1  # Adjust to 0-based index

    def cell_index(self,excel_index):

        match = re.match(r"^([A-Z]+)(\d+)$", excel_index)
        if not match:
            raise ValueError("Invalid index")
        
        x_cell = self.column_index(match.group(1))
        y_cell = int(match.group(2))-2
        
        return y_cell, x_cell

    def selection(self,cell= None, column= None, row= None, sheet= None, workbook= None):
        self.df_idx = []

        if cell:
            for i in cell:
                row_index,column_index= self.cell_index(i)
                self.df_idx.append((row_index,column_index))
                
        
        if column:
            for c in column:
                slice_variable = None
                column_index = self.column_index(c)
                slice_variable = slice(None), slice(column_index,column_index+1)
                self.df_idx.append(slice_variable)

        if row:
            for r in row:
                row_index = int(r-2)
                slice_variable = slice(row_index,row_index+1), slice(None)
                self.df_idx.append(slice_variable)

        if sheet:
            slice_variable = slice(None),slice(None)
            self.df_idx.append(slice_variable)

        if workbook:
            self.df_idx = [(':', ':') for _ in self.sheet_names]

                
    def dataframe(self,type,prompt=None):
        translated_df = self.df.copy()
        new_df_list = []
        new_df = None
        axis = 0 if type == "row" else 1

        for idx in self.df_idx:
            
            if type == "cell":
                
                if prompt:
                    translated_result = self.translate(self.df.iloc[idx],prompt)
                else:
                    translated_result = self.translate(self.df.iloc[idx])

                original.append(self.df.iloc[idx]) 
                translated.append(translated_result)

            elif type == "column" or type == "row":
                
                if prompt:
                    translated_result = self.df.iloc[idx].map(lambda cell_value:self.translate(cell_value,prompt))
                else:
                    translated_result = self.df.iloc[idx].map(self.translate)
                if type !="row":
                    translated_result.columns = [f"translated_{col}" for col in translated_result.columns]
                combined_data = pd.concat([self.df.iloc[idx], translated_result], axis=axis)
                new_df_list.append(combined_data)

            elif type == "sheet":
                if prompt:
                    translated_result = self.df.iloc[idx].map(lambda cell_value:self.translate(cell_value,prompt))
                else:
                    translated_result = self.df.iloc[idx].map(self.translate)

            translated_df.iloc[idx] = translated_result

        if type == "cell":

            ddf['Original'] = original
            ddf['Translated'] = translated
            new_df = ddf
            
        elif type == "row" or type == "column":
            new_df = pd.concat(new_df_list, axis=0)
            
        return new_df,self.df,translated_df
    

param_mapping = {
    "cell": "cell",
    "column": "column",
    "row": "row",
    "sheet": "sheet",
    "workbook": "sheet"
}

ddf = pd.DataFrame(columns=["Original","Translated"])


t=translator()

def read_sheets(file):

    workbook = pd.ExcelFile(file)
    sheet_names = workbook.sheet_names
    return sheet_names

def range_used(file,selected_sheet):

    wb = openpyxl.load_workbook(file)
    max_row = wb[selected_sheet].max_row
    max_column = wb[selected_sheet].max_column

    cell_list = [f"{openpyxl.utils.get_column_letter(col)}{row}" for col in range(1, max_column + 1) for row in range(2, max_row + 1)]
    # Generate the row list
    row_list = list(range(2, max_row + 1))

    # Generate the column list
    column_list = [openpyxl.utils.get_column_letter(col) for col in range(1, max_column + 1)]
    return cell_list,row_list,column_list

def process(file,selected_sheet,selection_input,type,prompt):

    t.excel_to_df(file= file,sheet_name=selected_sheet)
    kwargs = {param_mapping[type]: selection_input}
    t.selection(**kwargs)

    preview_df,original_df,translated_df=t.dataframe(param_mapping[type],prompt)

    if type=="sheet" or type=="workbook":
        return original_df,translated_df
    else:
        return preview_df,translated_df

